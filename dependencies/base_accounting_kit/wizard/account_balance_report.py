# -*- coding: utf-8 -*-
#############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2024-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions(<https://www.cybrosys.com>)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################
import io
import base64
from datetime import datetime
from odoo import api, fields, models
from odoo.exceptions import UserError
from odoo.tools.misc import get_lang
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class AccountBalanceReport(models.TransientModel):
    _name = 'account.balance.report'
    _inherit = "account.common.account.report"
    _description = 'Trial Balance Report'

    section_report_ids = fields.Many2many(string="Sections",
                                          comodel_name='account.report',
                                          relation="account_balance_report_section_rel",
                                          column1="main_report_id",
                                          column2="sub_report_id")
    section_main_report_ids = fields.Many2many(string="Section Of",
                                               comodel_name='account.report',
                                               relation="account_balance_report_section_rel",
                                               column1="sub_report_id",
                                               column2="main_report_id")
    name = fields.Char(string="Trial Balance", default="Trial Balance", required=True, translate=True)
    journal_ids = fields.Many2many('account.journal',
                                   'account_balance_report_journal_rel',
                                   'account_id', 'journal_id',
                                   string='Journals', required=True,
                                   default=[])

    @api.model
    def _get_report_name(self):
        period_id = self._get_selected_period_id()
        return self.env['consolidation.period'].browse(period_id)['display_name'] or self.env._("Trial Balance")

    def _print_report(self, data):
        data = self.pre_print_report(data)
        records = self.env[data['model']].browse(data.get('ids', []))
        return self.env.ref(
            'base_accounting_kit.action_report_trial_balance').report_action(
            records, data=data)

    def export_to_excel(self):
        """Export Trial Balance Report to Excel"""
        if not openpyxl:
            raise UserError(self.env._("Please install openpyxl library to export Excel files."))
        
        self.ensure_one()
        data = {}
        data['ids'] = self.env.context.get('active_ids', [])
        data['model'] = self.env.context.get('active_model', 'ir.ui.menu')
        data['form'] = self.read(['date_from', 'date_to', 'journal_ids', 'target_move',
                                  'display_account', 'company_id'])[0]
        used_context = self._build_contexts(data)
        data['form']['used_context'] = dict(used_context, lang='ar_001')
        data = self.pre_print_report(data)
        
        # Get account data directly using the report model's _get_accounts method
        # This avoids context issues with _get_report_values
        report_model = self.env['report.base_accounting_kit.report_trial_balance']
        model = data['model']
        display_account = data['form'].get('display_account', 'movement')
        
        # Get accounts based on model (same logic as in _get_report_values)
        if model == 'account.account' and data['ids']:
            accounts = self.env[model].browse(data['ids'])
        else:
            accounts = self.env['account.account'].search([])
        
        # Get account data with proper context
        account_res = report_model.with_context(
            data['form'].get('used_context', {})
        )._get_accounts(accounts, display_account)
        
        # Calculate initial balance for each account (balance before date_from)
        initial_balances = {}
        date_from = data['form'].get('date_from')
        if date_from:
            # Build context for initial balance calculation (before date_from)
            init_context = dict(data['form'].get('used_context', {}))
            init_context['date_from'] = False
            init_context['date_to'] = date_from
            init_context['strict_range'] = True
            
            # Get initial balances using the same method but with modified date range
            init_account_res = report_model.with_context(init_context)._get_accounts(accounts, 'all')
            # Create a map of code to balance for initial balances
            code_to_balance = {acc_data.get('code'): acc_data.get('balance', 0.0) for acc_data in init_account_res}
        else:
            code_to_balance = {}
        
        # Add initial balance to account_res
        for account_data in account_res:
            account_code = account_data.get('code')
            account_data['initial_balance'] = code_to_balance.get(account_code, 0.0)
    
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.env._("Trial Balance")
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        # Company name and title
        row = 1
        worksheet.merge_cells(f'A{row}:F{row}')
        cell = worksheet.cell(row=row, column=1)
        cell.value = f"{self.env.company.name}: {self.env._('Trial Balance')}"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Report parameters
        worksheet.cell(row=row, column=1, value=self.env._("Display Account:"))
        display_account_map = {
            'all': self.env._('All accounts'),
            'movement': self.env._('With movements'),
            'not_zero': self.env._('With balance not equal to zero')
        }
        worksheet.cell(row=row, column=2, value=display_account_map.get(data['form'].get('display_account', 'all'), self.env._('All accounts')))
        row += 1
        
        if data['form'].get('date_from'):
            worksheet.cell(row=row, column=1, value=self.env._("Date from:"))
            worksheet.cell(row=row, column=2, value=data['form']['date_from'])
            row += 1
        
        if data['form'].get('date_to'):
            worksheet.cell(row=row, column=1, value=self.env._("Date to:"))
            worksheet.cell(row=row, column=2, value=data['form']['date_to'])
            row += 1
        
        target_move_map = {
            'all': self.env._('All Entries'),
            'posted': self.env._('All Posted Entries')
        }
        worksheet.cell(row=row, column=1, value=self.env._("Target Moves:"))
        worksheet.cell(row=row, column=2, value=target_move_map.get(data['form'].get('target_move', 'all'), self.env._('All Entries')))
        row += 2
        
        # Table headers
        headers = [self.env._('Code'), self.env._('Account'), self.env._('Initial Balance'), self.env._('Debit'), self.env._('Credit'), self.env._('Balance')]
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1
        
        # Account data
        currency = self.env.company.currency_id
        accounts = account_res
        total_initial_balance = 0.0
        total_debit = 0.0
        total_credit = 0.0
        total_balance = 0.0
        
        for account in accounts:
            worksheet.cell(row=row, column=1, value=account.get('code', ''))
            worksheet.cell(row=row, column=2, value=account.get('name', ''))
            
            initial_balance = account.get('initial_balance', 0.0)
            debit = account.get('debit', 0.0)
            credit = account.get('credit', 0.0)
            balance = account.get('balance', 0.0)
            
            worksheet.cell(row=row, column=3, value=initial_balance)
            worksheet.cell(row=row, column=4, value=debit)
            worksheet.cell(row=row, column=5, value=credit)
            worksheet.cell(row=row, column=6, value=balance)
            
            # Format numbers
            for col in [3, 4, 5, 6]:
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            total_initial_balance += initial_balance
            total_debit += debit
            total_credit += credit
            total_balance += balance
            row += 1
        
        # Totals row
        row += 1
        worksheet.cell(row=row, column=1, value=self.env._("TOTAL"))
        worksheet.cell(row=row, column=2, value="")
        worksheet.cell(row=row, column=3, value=total_initial_balance)
        worksheet.cell(row=row, column=4, value=total_debit)
        worksheet.cell(row=row, column=5, value=total_credit)
        worksheet.cell(row=row, column=6, value=total_balance)
        
        # Style totals row
        totals_font = Font(bold=True)
        totals_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for col in range(1, 7):
            cell = worksheet.cell(row=row, column=col)
            cell.font = totals_font
            if col >= 3:
                cell.fill = totals_fill
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 18
        worksheet.column_dimensions['E'].width = 18
        worksheet.column_dimensions['F'].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Encode to base64
        excel_file = base64.b64encode(output.read())
        output.close()
        
        # Generate filename
        filename = f"Trial_Balance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Return download action
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=ir.attachment&field=datas&filename_field=name&id=%s&download=true' % (
                self.env['ir.attachment'].create({
                    'name': filename,
                    'datas': excel_file,
                    'type': 'binary',
                    'res_model': self._name,
                    'res_id': self.id,
                }).id
            ),
            'target': 'self',
        }
